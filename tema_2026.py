# -*- coding: utf-8 -*-
"""
Оптимизированная версия инструмента реестра отклонений ITSM и АС.
Точка входа: CLI с подкомандами create-forms | build-registry | quarterly-increment.
Совместимость выходных Excel с tema_2025.py сохранена.
"""
from __future__ import annotations

import argparse
import logging
import os
import re
import subprocess
import sys
import warnings
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from dateutil import parser

# Опционально: sheet2dict для режимов 1/2 при отказе от subprocess
try:
    from sheet2dict import Worksheet
except ImportError:
    Worksheet = None

# -----------------------------------------------------------------------------
# Конфигурация
# -----------------------------------------------------------------------------

@dataclass
class Config:
    """Конфигурация путей и параметров (можно переопределить через CLI/переменные окружения)."""
    script_dir: str = field(default_factory=lambda: os.path.abspath(os.path.dirname(__file__)))
    folder_no_hand: str = "no_hand"
    folder_as: str = "AS"
    # Имена файлов (подстрока для поиска по glob)
    keyword_oprosnik: str = "Опросник"
    keyword_baza_znaniy: str = "База_знаний"
    keyword_anketa_objects: str = "Анкета_объектов"
    keyword_edk_ero: str = "ЕДК"  # или ЕРО
    expected_oprosnik_version_prefix: str = "6.1."
    template_anketa: str = "Анкета - ИТ-услуга.xlsx"
    template_anketa_script: str = "Анкета - ИТ-для скрипта.xlsx"
    sheet_reestr: str = "Реестр объектов обследования"
    sheet_harakteristiki: str = "Характеристики ИТ-услуги"
    sheet_nadezhnost: str = "Надёжность"
    sheet_baza: str = "База"
    sheet_baza_as: str = "База АС"
    sheet_edk: str = "Единая дорожная карта"

    def path_no_hand(self) -> str:
        return os.path.join(self.script_dir, self.folder_no_hand)

    def path_as(self) -> str:
        return os.path.join(self.script_dir, self.folder_as)


# -----------------------------------------------------------------------------
# Логирование
# -----------------------------------------------------------------------------

def setup_logging(verbose: bool = False) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(message)s",
        stream=sys.stdout,
    )


log = logging.getLogger(__name__)


# -----------------------------------------------------------------------------
# Даты и кварталы (общий модуль)
# -----------------------------------------------------------------------------

MONTHS = {
    'янв': 1, 'фев': 2, 'мар': 3, 'апр': 4, 'май': 5, 'июн': 6,
    'июл': 7, 'авг': 8, 'сен': 9, 'окт': 10, 'ноя': 11, 'дек': 12,
    'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12,
}


def get_current_quarter() -> str:
    """Текущий квартал в формате nQYYYY."""
    now = datetime.now()
    q = (now.month - 1) // 3 + 1
    return f"{q}Q{now.year}"


def extract_quarter_year(q_str: str) -> tuple[int, int]:
    """Извлекает квартал и год из строки формата 'nQYYYY' или 'QnYYYY'."""
    s = q_str.strip().upper()
    if s.startswith("Q"):
        quarter = int(s[1])
        year = int(s[2:])
    else:
        parts = s.split("Q")
        if len(parts) != 2:
            raise ValueError(f"Некорректный формат квартала: {q_str}")
        quarter = int(parts[0])
        year = int(parts[1])
    return quarter, year


def find_future_quarters(current_q: str, all_quarters: list[str]) -> list[str]:
    """Кварталы из all_quarters, строго большие текущего."""
    try:
        cq, cy = extract_quarter_year(current_q)
        out = []
        for q in all_quarters:
            try:
                qq, qy = extract_quarter_year(q)
                if qy > cy or (qy == cy and qq > cq):
                    out.append(q)
            except ValueError:
                log.warning("Некорректная дата в списке: %s", q)
                continue
        return sorted(out, key=lambda x: (extract_quarter_year(x)))
    except ValueError as e:
        log.warning("Ошибка текущего квартала '%s': %s", current_q, e)
        return []


def date_to_quarter_year(date_str: str) -> Optional[str]:
    """Парсит дату/квартал в формат nQYYYY. Поддержка: 1кв.2025, Q1 25, июн.2025, 31.03.2025, 2025-07-31."""
    if not date_str or not isinstance(date_str, str):
        return None
    s = date_str.strip().lower()

    # Квартальные форматы
    qm = (
        re.match(r'(?:q|кв)(\d)[\'\s\.]*(\d{2,4})', s) or
        re.match(r'(\d)\s*(?:кв|q)[\'\s\.]*(\d{2,4})', s) or
        re.match(r'(\d)\s*квартал\s*(\d{4})', s) or
        re.match(r'(\d{4})\s*(?:q|кв)\s*(\d)', s)
    )
    if qm:
        if '(\d{4})' in qm.re.pattern and qm.re.pattern.startswith(r'(\d{4})'):
            year = int(qm.group(1))
            quarter = int(qm.group(2))
        else:
            quarter = int(qm.group(1))
            year = int(qm.group(2))
        return f"{quarter}Q{year if year >= 100 else year + 2000}"

    # Текстовый месяц
    mm = (
        re.match(r'([а-яa-z]+)\.?\s*(\d{4})', s) or
        re.match(r'(\d{1,2})-([а-яa-z]+)-(\d{4})', s) or
        re.match(r'(\d{1,2})\s+([а-яa-z]+)\s+(\d{4})', s)
    )
    if mm:
        g = mm.groups()
        if len(g) == 2:
            month_str = g[0][:3]
            year = int(g[1])
        else:
            month_str = g[1][:3]
            year = int(g[2])
        if month_str in MONTHS:
            quarter = (MONTHS[month_str] - 1) // 3 + 1
            return f"{quarter}Q{year}"

    try:
        parsed = parser.parse(s, dayfirst=True)
        q = (parsed.month - 1) // 3 + 1
        return f"{q}Q{parsed.year}"
    except (ValueError, TypeError):
        return None


def add_quarters(initial_date: str, quarters_to_add: int) -> str:
    """Добавляет quarters_to_add кварталов к initial_date (формат nQYYYY или QnYYYY)."""
    s = initial_date.strip().upper()
    if s.startswith("Q"):
        quarter = int(s[1])
        year = int(s[2:])
    else:
        parts = s.split("Q")
        quarter = int(parts[0])
        year = int(parts[1])
    total = quarter + quarters_to_add - 1
    new_year = year + total // 4
    new_quarter = (total % 4) + 1
    return f"{new_quarter}Q{new_year}"


def find_min_quarter(result_mas: list) -> Optional[str]:
    """Минимальный квартал в списке строк формата nQYYYY."""
    min_year = float('inf')
    min_quarter = float('inf')
    out = None
    for q in result_mas:
        if q is None:
            continue
        try:
            if 'Q' not in q:
                continue
            parts = q.split('Q')
            if len(parts) == 2:
                qu, yr = int(parts[0]), int(parts[1])
            else:
                qu, yr = int(q[1]), int(q[2:])
            if yr < min_year or (yr == min_year and qu < min_quarter):
                min_year, min_quarter = yr, qu
                out = f"{qu}Q{yr}"
        except (ValueError, IndexError, AttributeError):
            continue
    return out


# -----------------------------------------------------------------------------
# Справочники
# -----------------------------------------------------------------------------

DICT_INN_COMPANY: dict[str, list[str]] = {
    "9705118142": ["Подразделения вне блоков B2C", "Купер"],
    "9701048328": ["Подразделения вне блоков B2C", "МегаМаркет"],
    "7811554010": ["Подразделения вне блоков B2C", "Самокат"],
    "7736322345": ["Подразделения вне блоков B2C", "СберЛогистика"],
    "7730262964": ["EdTech", "СберОбразование"],
    "7736316133": ["EdTech", "Школа 21"],
    "7704865540": ["E-Health", "ЕАптека"],
    "9710011437": ["E-Health", "СберЗдоровье"],
    "9731065465": ["E-Health", "СберМедИИ"],
    "9705124940": ["GR", "СберПраво"],
    "7708328948": ["Media & Ads", "Звук"],
    "7801445445": ["Media & Ads", "Звук Бизнес"],
    "7814665871": ["Media & Ads", "ОККО"],
    "7725243282": ["Media & Ads", "Рамблер"],
    "7736319695": ["Media & Ads", "СберМаркетинг"],
    "7736659589": ["ДРПА", "АктивБизнесКонсалт"],
    "7736303529": ["ДРПА", "АктивБизнесТехнологии"],
    "7736581290": ["ДРПА", "Сбербанк Капитал"],
    "5405276278": ["КИБ", "2ГИС"],
    "7736641983": ["КИБ", "Деловая среда"],
    "9731062087": ["КИБ", "Дома"],
    "7714843760": ["КИБ", "Инсейлс"],
    "7709969870": ["КИБ", "Работа.ру"],
    "7707308480": ["КИБ", "Сбер А"],
    "7730269550": ["КИБ", "Сбер Бизнес Софт"],
    "7730241227": ["КИБ", "СберАналитика"],
    "7707009586": ["КИБ", "Сбербанк Лизинг"],
    "7802754982": ["КИБ", "Сбербанк Факторинг"],
    "7801392271": ["КИБ", "СберКорус"],
    "7709688816": ["КИБ", "СберРешения"],
    "7730262971": ["КИБ", "СберТаксФри"],
    "7736612855": ["КИБ", "Стратеджи Партнерс Групп"],
    "7727381792": ["КИБ", "Фьюэл-Ап"],
    "7730261382": ["КИБ", "Цифровые решения регионов"],
    "9709108748": ["ЛК", "Пульс"],
    "7736128605": ["ЛК", "СберУниверситет"],
    "7736249247": ["Подразделения вне блоков B2C", "ДомКлик"],
    "9709054813": ["Подразделения вне блоков B2C", "СберАвто"],
    "7736264044": ["Подразделения вне блоков B2C", "СберМобайл"],
    "7704314221": ["Подразделения вне блоков B2C", "Ситидрайв"],
    "9709078370": ["Подразделения вне блоков B2C", "Центр новых финансовых сервисов"],
    "7702770003": ["Развитие клиентского опыта B2C", "СберСпасибо"],
    "7710561081": ["Риски", "ОКБ"],
    "7736324991": ["Сервисы", "БАРУС"],
    "400014449": ["Сервисы", "Манжерок"],
    "7730245060": ["Сервисы", "Медэксперт Плюс"],
    "7729276546": ["Сервисы", "Московский городской Гольф Клуб"],
    "9103007830": ["Сервисы", "Мрия"],
    "7736663049": ["Сервисы", "СберСервис"],
    "7708229993": ["Сервисы", "СовТех"],
    "7720427871": ["Сервисы", "СТК"],
    "9709073460": ["Сеть продаж", "СберМегаМаркетРитейл"],
    "5024093941": ["Строительство", "Рублево-Архангельское"],
    "9731026963": ["Строительство", "Смарт Констракшн"],
    "6439098794": ["Строительство", "Инфотех Балаково"],
    "7736632467": ["Технологии", "Сбербанк-Технологии"],
    "7736279160": ["Технологическое развитие", "Cloud.ru"],
    "9725045830": ["Технологическое развитие", "АВТОТЕХ"],
    "7725745476": ["Технологическое развитие", "Живой Сайт"],
    "7730253720": ["Технологическое развитие", "СалютДевайсы"],
    "7805093681": ["Технологическое развитие", "ЦРТ"],
    "7727718421": ["Транзакционный банкинг B2C", "Расчетные решения"],
    "7750005860": ["Транзакционный банкинг B2C", "Расчетные решения НКО"],
    "9702027017": ["Транзакционный банкинг B2C", "СберТройка"],
    "9715225506": ["Транзакционный банкинг B2C", "Эвотор"],
    "7750005725": ["Транзакционный банкинг B2C", "Юмани"],
    "7725352740": ["УБ", "НПФ Сбербанка"],
    "9725000621": ["УБ", "Пенсионные решения"],
    "7736618039": ["УБ", "Рыночный Спецдепозитарий"],
    "7744002123": ["УБ", "Сбербанк страхование жизни СК"],
    "7706810747": ["УБ", "Сбербанк страхование СК"],
    "7730257675": ["УБ", "Современные Фонды Недвижимости"],
    "7706810730": ["УБ", "Страховой брокер Сбербанка"],
    "7710183778": ["УБ", "Управляющая компания Первая"],
    "7736252313": ["ЦПНД", "Цифровые технологии"],
}

CRITICALS_ENG_TO_CODE = {
    "Office Productivity": "OP",
    "Business Operational": "BO",
    "Business Critical": "BC",
    "Mission Critical": "MC",
}

VALID_CRITICALITY_OPROSNIK = ["Нет подобных ИТ-сервисов", "OP", "BO", "BC и выше"]


# -----------------------------------------------------------------------------
# Режим 0: Создание анкет ИТ-услуг
# -----------------------------------------------------------------------------

def _find_code_column(sheet) -> Optional[tuple[str, int]]:
    """Находит столбец с заголовком 'Код' и номер строки. Возвращает (буква_столбца, номер_строки)."""
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and "Код" in str(cell.value):
                from openpyxl.utils import get_column_letter
                return get_column_letter(cell.column), cell.row
    return None


def _load_app_services_from_reestr(
    path_anketa_objects: str,
    sheet_name: str,
) -> dict:
    """Загружает реестр и возвращает словарь {код: [название, критичность_eng, описание], ...} только для типа 'Прикладной сервис'."""
    wb = openpyxl.load_workbook(path_anketa_objects, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise FileNotFoundError(f"Лист '{sheet_name}' не найден в файле {path_anketa_objects}")
    sheet = wb[sheet_name]
    code_col_info = _find_code_column(sheet)
    if not code_col_info:
        raise ValueError("В файле не найдена колонка с заголовком 'Код'")
    header_row = code_col_info[1]
    headers = []
    for c in range(1, sheet.max_column + 1):
        v = sheet.cell(header_row, c).value
        headers.append((c, str(v).strip() if v else ""))
    col_code = next((c for c, h in headers if h == "Код"), None)
    col_type = next((c for c, h in headers if "Тип" in h or h == "Тип объекта"), None)
    col_name = next((c for c, h in headers if "Название" in h or h == "Название"), None)
    col_crit = next((c for c, h in headers if "Критичность" in h), None)
    col_desc = next((c for c, h in headers if "Описание" in h), None)
    if not col_code or not col_type:
        raise ValueError("Не найдены колонки Код или Тип объекта")
    result = {}
    for r in range(header_row + 1, sheet.max_row + 1):
        typ = sheet.cell(r, col_type).value
        if not typ or str(typ).strip() != "Прикладной сервис":
            continue
        code = sheet.cell(r, col_code).value
        if not code:
            continue
        name = (sheet.cell(r, col_name).value or "") if col_name else ""
        crit = (sheet.cell(r, col_crit).value or "") if col_crit else ""
        desc = (sheet.cell(r, col_desc).value or "") if col_desc else ""
        result[str(code).strip()] = [str(name).strip(), str(crit).strip(), str(desc).strip()]
    wb.close()
    return result


def _copy_validations_from_template(ws_target, ws_template, cell_addresses: set, formula_contains: str = "Справочники") -> None:
    """Копирует валидации списка из ws_template в ws_target для ячеек из cell_addresses."""
    try:
        validations = list(ws_template.data_validations.dataValidation)
    except Exception:
        return
    validation_by_cell = {}
    for v in validations:
        if formula_contains not in (v.formula1 or ""):
            continue
        for cell_ref in v.cells:
            validation_by_cell[str(cell_ref)] = v
    expanded = {}
    for ref, v in validation_by_cell.items():
        if ":" in ref:
            a, b = ref.split(":")
            c1, r1 = a[0], int(a[1:])
            c2, r2 = b[0], int(b[1:])
            for ri in range(r1, r2 + 1):
                expanded[f"{c1}{ri}"] = v
        else:
            expanded[ref] = v
    for addr in cell_addresses:
        if addr in expanded:
            f1 = (expanded[addr].formula1 or "").replace('"', '')
            try:
                dv = DataValidation(type="list", formula1=f1, allow_blank=True)
                dv.add(ws_target[addr])
                ws_target.add_data_validation(dv)
            except Exception:
                pass


def mode_0_create_forms(
    config: Config,
    company: str,
    block: str,
    expert_dtn: str,
    expert_dka: str,
) -> None:
    """Режим 0: создание анкет ИТ-услуг в папке AS."""
    warnings.simplefilter("ignore")
    no_hand = config.path_no_hand()
    as_path = config.path_as()
    script_dir = config.script_dir

    if not os.path.isdir(no_hand):
        log.error("Папка %s не найдена.", no_hand)
        sys.exit(1)

    files = [f for f in os.listdir(no_hand) if f.endswith(".xlsx") and config.keyword_anketa_objects in f]
    if not files:
        log.error("В папке %s не найден файл с '%s'. Анкеты не будут сформированы.", no_hand, config.keyword_anketa_objects)
        sys.exit(1)

    path_anketa_objects = os.path.join(no_hand, files[0])
    log.info("Используется файл реестра: %s", path_anketa_objects)

    dtn_as_all = _load_app_services_from_reestr(
        path_anketa_objects,
        config.sheet_reestr,
    )
    if not dtn_as_all:
        log.warning("Нет записей типа 'Прикладной сервис' в реестре.")
        return

    path_template = os.path.join(no_hand, config.template_anketa)
    path_template_script = os.path.join(no_hand, config.template_anketa_script)
    if not os.path.isfile(path_template):
        log.error("Шаблон не найден: %s", path_template)
        sys.exit(1)
    if not os.path.isfile(path_template_script):
        log.error("Шаблон для валидаций не найден: %s", path_template_script)
        sys.exit(1)

    os.makedirs(as_path, exist_ok=True)
    wb_two = openpyxl.load_workbook(path_template_script)
    ws_techno_two = wb_two["Технологии"]
    ws_tech_two = wb_two["Компоненты ИТ-услуги"]
    cells_techno = {f"C{i}" for i in range(1, 115)}
    cells_tech_a = {f"A{i}" for i in range(1, 25)}
    cells_tech_l = {f"L{i}" for i in range(1, 25)}

    for code, values in dtn_as_all.items():
        name_svc, crit_eng, desc = values[0], values[1], values[2]
        wb = openpyxl.load_workbook(path_template)
        ws = wb[config.sheet_harakteristiki]
        ws_techno = wb["Технологии"]
        ws_tech = wb["Компоненты ИТ-услуги"]

        ws["C2"].value = company
        ws["C4"].value = datetime.now().year
        ws["C5"].value = block
        ws["C18"].value = expert_dtn
        ws["C19"].value = expert_dka
        ws["C3"].value = name_svc
        ws["C6"].value = code
        ws["C7"].value = CRITICALS_ENG_TO_CODE.get(crit_eng, crit_eng)
        ws["C9"].value = desc

        _copy_validations_from_template(ws_techno, ws_techno_two, cells_techno)
        _copy_validations_from_template(ws_tech, ws_tech_two, cells_tech_a)
        _copy_validations_from_template(ws_tech, ws_tech_two, cells_tech_l)

        safe_name = re.sub(r'[\\/:*?"<>|]', '', name_svc)
        out_path = os.path.join(as_path, f"Анкета - ИТ-услуга-{safe_name}.xlsx")
        wb.save(out_path)
        wb.close()
        log.info("Создана анкета: %s", out_path)

    wb_two.close()
    log.info("Режим 0 завершён. Создано анкет: %d", len(dtn_as_all))


# -----------------------------------------------------------------------------
# Режимы 1 и 2: делегирование tema_2025.py для полной совместимости
# -----------------------------------------------------------------------------

def _run_tema_2025_mode(mode: int, overwrite: bool) -> int:
    """Запускает tema_2025.py с выбранным режимом (1 или 2). Передаёт overwrite через stdin при необходимости."""
    script_2025 = os.path.join(os.path.dirname(__file__), "tema_2025.py")
    if not os.path.isfile(script_2025):
        log.error("Файл tema_2025.py не найден. Режимы 1 и 2 требуют его наличия.")
        return 1
    stdin_input = str(mode)
    if overwrite:
        stdin_input += "\ny"  # ответ перезаписи при запросе
    try:
        proc = subprocess.run(
            [sys.executable, script_2025],
            input=stdin_input.encode("utf-8"),
            cwd=os.path.dirname(script_2025),
            timeout=600,
        )
        return proc.returncode
    except subprocess.TimeoutExpired:
        log.error("Таймаут выполнения tema_2025.py")
        return 1
    except Exception as e:
        log.error("Ошибка запуска tema_2025.py: %s", e)
        return 1


def mode_1_build_registry(config: Config, overwrite: bool = False) -> int:
    """Режим 1: генерация реестра ITSM + АС. Делегирует tema_2025.py для полной совместимости."""
    log.info("Запуск генерации реестра (режим 1) через tema_2025.py...")
    return _run_tema_2025_mode(1, overwrite)


def mode_2_quarterly_increment(config: Config, overwrite: bool = False) -> int:
    """Режим 2: ежеквартальный инкремент после ЕОИТ в ДЗО. Делегирует tema_2025.py."""
    log.info("Запуск расчёта квартального инкремента (режим 2) через tema_2025.py...")
    return _run_tema_2025_mode(2, overwrite)


# -----------------------------------------------------------------------------
# CLI
# -----------------------------------------------------------------------------

def _print_banner() -> None:
    print("=" * 100)
    print("Инструмент реестра отклонений ITSM и АС (оптимизированная версия).")
    print("=" * 100)
    print("Режимы:")
    print("  create-forms         — создание анкет ИТ-услуг для оценки (режим 0)")
    print("  build-registry       — генерация реестра ITSM + АС (режим 1)")
    print("  quarterly-increment  — расчёт ежеквартального инкремента после ЕОИТ в ДЗО (режим 2)")
    print("=" * 100)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Реестр отклонений ITSM и АС. Режимы: create-forms | build-registry | quarterly-increment.",
    )
    parser.add_argument(
        "command",
        choices=["create-forms", "build-registry", "quarterly-increment"],
        help="Режим работы",
    )
    parser.add_argument("--company", default="", help="Название компании (режим 0)")
    parser.add_argument("--block", default="", help="Блок (режим 0)")
    parser.add_argument("--expert-dtn", default="", help="Эксперт по надёжности ДТН (режим 0)")
    parser.add_argument("--expert-dka", default="", help="Куратор по архитектуре ДКА (режим 0)")
    parser.add_argument("--overwrite", action="store_true", help="Перезаписывать выходные файлы без запроса")
    parser.add_argument("--verbose", "-v", action="store_true", help="Подробный вывод")
    parser.add_argument("--config-dir", default="", help="Рабочая директория (по умолчанию — каталог скрипта)")

    args = parser.parse_args()
    setup_logging(args.verbose)

    if args.config_dir:
        os.chdir(args.config_dir)

    config = Config()
    if args.config_dir:
        config.script_dir = os.path.abspath(args.config_dir)

    if args.command == "create-forms":
        _print_banner()
        company = args.company or input("Введите наименование компании: ").strip()
        block = args.block or input("Введите название блока: ").strip()
        expert_dtn = args.expert_dtn or input("Эксперт по надёжности (ДТН): ").strip()
        expert_dka = args.expert_dka or input("Куратор по архитектуре (ДКА): ").strip()
        mode_0_create_forms(config, company, block, expert_dtn, expert_dka)
        return 0

    if args.command == "build-registry":
        _print_banner()
        return mode_1_build_registry(config, overwrite=args.overwrite)

    if args.command == "quarterly-increment":
        _print_banner()
        return mode_2_quarterly_increment(config, overwrite=args.overwrite)

    return 0


if __name__ == "__main__":
    sys.exit(main() or 0)
